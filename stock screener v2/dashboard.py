import sys
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import os
import threading
import time
import datetime
import openpyxl
import yfinance as yf

# ==========================================
# 1. PAGE CONFIG & UI STYLING
# ==========================================
st.set_page_config(page_title="V2 Trading Dashboard", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    div[data-testid="stMetricValue"] {
        font-size: 2rem !important; font-weight: 700; color: #00F0FF;
        text-shadow: 0px 0px 10px rgba(0, 240, 255, 0.4);
    }
    div[data-testid="stMetricLabel"] { font-size: 1rem !important; font-weight: 600; color: #B0BEC5; }
    .dataframe { border-radius: 10px; overflow: hidden; }
    h1 { font-weight: 700 !important; background: -webkit-linear-gradient(45deg, #00F0FF, #00FF88);
         -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
    h2, h3 { color: #FFFFFF; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

V2_DIR = os.path.dirname(os.path.abspath(__file__))
TRACKER_FILE = os.path.join(V2_DIR, "tracker.xlsx")
POLL_INTERVAL = 60 # seconds

# ==========================================
# 2. PAPER TRADING ENGINE (BACKGROUND)
# ==========================================
MARKUP_PCT = 0.0002  # 0.02% execution markup / slippage

def fetch_live_prices(tickers):
    """Fetch recent 1m data for given tickers to get Open/High/Low bounds"""
    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] Fetching 1m data for {len(tickers)} tickers...")
    try:
        data = yf.download(tickers, period="2d", interval="1m", progress=False)
        prices = {}
        if len(tickers) == 1:
            sym = tickers[0]
            if not data.empty:
                open_val = data['Open'].dropna().iloc[0] if ('Open' in data and not data['Open'].dropna().empty) else data['Close'].iloc[-1]
                prices[sym] = {
                    'Open': float(open_val),
                    'High': float(data['High'].max()),
                    'Low': float(data['Low'].min()),
                    'LTP': float(data['Close'].iloc[-1])
                }
        else:
            for sym in tickers:
                if sym in data['Close']:
                    high = data['High'][sym].max()
                    low = data['Low'][sym].min()
                    ltp = data['Close'][sym].dropna().iloc[-1] if not data['Close'][sym].dropna().empty else None
                    open_val = data['Open'][sym].dropna().iloc[0] if ('Open' in data and sym in data['Open'] and not data['Open'][sym].dropna().empty) else ltp
                    if ltp is not None:
                        prices[sym] = {
                            'Open': float(open_val) if open_val is not None else float(ltp),
                            'High': float(high),
                            'Low': float(low),
                            'LTP': float(ltp)
                        }
        return prices
    except Exception as e:
        print(f"Error fetching data: {e}")
        return {}

def run_paper_trader():
    # Only poll live price candles during trading hours (8:30 AM to 4:00 PM IST on Weekdays)
    now = datetime.datetime.now()
    if not (now.weekday() < 5 and (8 <= now.hour < 16)):
        return

    if not os.path.exists(TRACKER_FILE):
        return
        
    try:
        wb = openpyxl.load_workbook(TRACKER_FILE)
    except PermissionError:
        print(f"[ERROR] Cannot open Excel file. Please close '{TRACKER_FILE}'.")
        return

    if 'Trade Log' not in wb.sheetnames:
        return
    ws = wb['Trade Log']
    
    active_rows = []
    tickers = set()
    
    for row_idx in range(2, ws.max_row + 1):
        stock = ws.cell(row=row_idx, column=2).value
        status = ws.cell(row=row_idx, column=18).value
        
        if stock and (status is None or status == "OPEN"):
            entry = ws.cell(row=row_idx, column=3).value
            sl = ws.cell(row=row_idx, column=4).value
            t1 = ws.cell(row=row_idx, column=5).value
            t2 = ws.cell(row=row_idx, column=6).value
            
            if entry and sl and t1 and t2:
                date_val = ws.cell(row=row_idx, column=1).value
                date_str = date_val.strftime("%Y-%m-%d") if isinstance(date_val, (datetime.date, datetime.datetime)) else str(date_val) if date_val else ""
                    
                active_rows.append({
                    'row': row_idx, 'stock': stock, 'status': status, 'date': date_str,
                    'entry': float(entry), 'sl': float(sl), 't1': float(t1), 't2': float(t2)
                })
                tickers.add(stock + ".NS")
                
    if not active_rows:
        return
        
    made_changes = False
    live_data = fetch_live_prices(list(tickers))
    
    for trade in active_rows:
        sym = trade['stock'] + ".NS"
        if sym not in live_data:
            continue
            
        high = live_data[sym]['High']
        low = live_data[sym]['Low']
        
        row_idx = trade['row']
        status = str(trade['status']) if trade['status'] is not None else ""
        
        # 1. Check for ENTRY if PENDING
        if not status or status == "None":
            today_str = datetime.date.today().strftime("%Y-%m-%d")
            if today_str > trade['date'] and high >= trade['entry']:
                # Calculate realistic execution price (handles gap-ups + 0.02% markup)
                open_price = live_data[sym].get('Open', trade['entry'])
                base_exec = max(open_price, trade['entry'])
                actual_exec = round(base_exec * (1.0 + MARKUP_PCT), 2)  # 0.02% markup
                
                print(f"[ENTRY TRIGGERED] {trade['stock']} crossed Entry {trade['entry']:.2f}. Executed at ₹{actual_exec:.2f} (Gap/0.02% Markup adjusted). Marking OPEN.")
                ws.cell(row=row_idx, column=3).value = actual_exec   # Update Entry Price in Excel
                ws.cell(row=row_idx, column=18).value = "OPEN"
                made_changes = True
                status = "OPEN" 

                try:
                    from telegram_bot import send_trade_event_alert
                    send_trade_event_alert("ENTRY", trade['stock'], actual_exec, f"SL: ₹{trade['sl']:.2f} | T1: ₹{trade['t1']:.2f} | T2: ₹{trade['t2']:.2f}")
                except Exception:
                    pass
                
        # 2. Check for EXIT if OPEN or OPEN - T1 HIT
        if status and "OPEN" in status:
            exit_hit, exit_price, exit_reason = False, 0.0, ""
            
            # Check SL / Breakeven SL with Gap-down protection + 0.03% exit slippage
            if low <= trade['sl']:
                exit_hit = True
                open_price = live_data[sym].get('Open', trade['sl'])
                base_exit = min(open_price, trade['sl'])
                exit_price = round(base_exit * (1.0 - 0.0003), 2)  # 0.03% exit slippage
                exit_reason = "CLOSED - SL" if status == "OPEN" else "CLOSED - BREAKEVEN SL"
                
            elif high >= trade['t2']:
                exit_hit = True
                exit_price = round(trade['t2'] * (1.0 - 0.0002), 2)
                exit_reason = "CLOSED - TARGET 2"
                
            elif high >= trade['t1'] and status == "OPEN":
                # PARTIAL PROFIT BOOKING AT TARGET 1 (50% Qty)
                t1_exit = round(trade['t1'] * (1.0 - 0.0002), 2)
                orig_qty = trade.get('qty', 0.0)
                half_qty = math.floor(orig_qty / 2.0) if orig_qty > 1 else orig_qty
                rem_qty = orig_qty - half_qty
                
                breakeven_sl = round(trade['entry'] * 1.0015, 2)
                
                print(f"[TARGET 1 HIT - PARTIAL 50%] {trade['stock']} hit T1 {trade['t1']:.2f}. Booking {half_qty} shares at ₹{t1_exit:.2f}, trailing {rem_qty} shares with Breakeven SL ₹{breakeven_sl:.2f}.")
                
                ws.cell(row=row_idx, column=4).value = breakeven_sl   # Column D: SL Price
                ws.cell(row=row_idx, column=11).value = rem_qty        # Column K: Qty
                ws.cell(row=row_idx, column=12).value = round(rem_qty * trade['entry'], 2) # Column L: Position Size
                ws.cell(row=row_idx, column=18).value = "OPEN - T1 HIT"
                made_changes = True
                
                try:
                    send_trade_event_alert("T1_PARTIAL", trade['stock'], t1_exit, f"₹{breakeven_sl:.2f}")
                except Exception:
                    pass
                
            if exit_hit:
                print(f"[{exit_reason}] {trade['stock']} exited at ₹{exit_price:.2f} (Gap/Slippage adjusted).")
                ws.cell(row=row_idx, column=18).value = exit_reason
                ws.cell(row=row_idx, column=19).value = exit_price
                made_changes = True
                
                try:
                    kind = "TARGET_2" if "TARGET 2" in exit_reason else "SL_EXIT"
                    send_trade_event_alert(kind, trade['stock'], exit_price, exit_reason)
                except Exception:
                    pass

    if made_changes:
        try:
            wb.save(TRACKER_FILE)
        except PermissionError:
            pass

# ==========================================
# 2.5 TELEGRAM BOT ENGINE (EMBEDDED IN DASHBOARD)
# ==========================================
import requests

DEFAULT_BOT_TOKEN = ""
DEFAULT_CHAT_ID   = ""

def get_telegram_config():
    bot_token = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
    chat_id   = os.environ.get("TELEGRAM_CHAT_ID", "").strip()

    try:
        if hasattr(st, "secrets"):
            if "TELEGRAM_BOT_TOKEN" in st.secrets:
                bot_token = str(st.secrets["TELEGRAM_BOT_TOKEN"]).strip()
            if "TELEGRAM_CHAT_ID" in st.secrets:
                chat_id = str(st.secrets["TELEGRAM_CHAT_ID"]).strip()
    except Exception:
        pass

    if (not bot_token or not chat_id) and os.path.exists(TRACKER_FILE):
        try:
            wb = openpyxl.load_workbook(TRACKER_FILE, data_only=True)
            if 'Config' in wb.sheetnames:
                ws = wb['Config']
                for row in ws.iter_rows(min_row=2, max_row=20, values_only=True):
                    if not row[0]: continue
                    param = str(row[0]).lower().strip()
                    if "bot_token" in param or "telegram token" in param or "bot token" in param:
                        if row[1] and str(row[1]).strip() and not str(row[1]).startswith("="):
                            bot_token = str(row[1]).strip()
                    elif "chat_id" in param or "telegram chat" in param or "chat id" in param:
                        if row[1] and str(row[1]).strip() and not str(row[1]).startswith("="):
                            chat_id = str(row[1]).strip()
        except Exception as e:
            print("Error reading Telegram config from Excel:", e)

    return bot_token or DEFAULT_BOT_TOKEN, chat_id or DEFAULT_CHAT_ID

def send_telegram_message(text, parse_mode="HTML"):
    bot_token, chat_id = get_telegram_config()
    if not bot_token or not chat_id:
        return False
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {"chat_id": chat_id, "text": text, "parse_mode": parse_mode, "disable_web_page_preview": True}
    try:
        r = requests.post(url, data=payload, timeout=10)
        return r.status_code == 200
    except Exception as e:
        print("[TELEGRAM ERROR]", e)
        return False

def send_telegram_message_with_keyboard(text, reply_markup, parse_mode="HTML"):
    bot_token, chat_id = get_telegram_config()
    if not bot_token or not chat_id:
        return False
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {"chat_id": chat_id, "text": text, "parse_mode": parse_mode, "reply_markup": reply_markup, "disable_web_page_preview": True}
    try:
        r = requests.post(url, json=payload, timeout=10)
        return r.status_code == 200
    except Exception as e:
        print("[TELEGRAM ERROR]", e)
        return False

def answer_callback_query(callback_query_id, text=""):
    bot_token, _ = get_telegram_config()
    url = f"https://api.telegram.org/bot{bot_token}/answerCallbackQuery"
    payload = {"callback_query_id": callback_query_id, "text": text}
    try:
        requests.post(url, json=payload, timeout=5)
    except Exception:
        pass

def send_daily_picks_alert(trades, watchlist):
    today_str = datetime.date.today().strftime('%d %b %Y')
    msg = f"📊 <b>V2 DAILY PICK SUMMARY ({today_str})</b>\n───────────────────────────\n\n"
    if trades:
        msg += f"🚀 <b>Top Qualified Trades for Tomorrow ({len(trades)}):</b>\n"
        for idx, t in enumerate(trades, 1):
            msg += f"<b>{idx}. {t['Stock']}</b> ({t.get('Sector', 'General')})\n   • Entry: ₹{t['Entry']:.2f}\n   • SL: ₹{t['SL']:.2f} ({t['SL %']:.1f}%)\n   • T1: ₹{t['T1']:.2f} | T2: ₹{t['T2']:.2f}\n   • Qty: {t['Qty']} shares | Size: ₹{t['Position Size']:,.0f}\n   • Momentum Score: {t['Score']:.2f}\n\n"
    else:
        msg += "⚠️ <i>No confirmed breakout trades qualified today.</i>\n\n"
    if watchlist:
        msg += f"👀 <b>Near-Breakout Watchlist ({len(watchlist)}):</b>\n"
        for w in watchlist[:5]:
            msg += f"   • <b>{w['Stock']}</b>: ₹{w['LTP']:.2f} ({w['% from High']:.1f}% below 52W High)\n"
    send_telegram_message(msg)

def send_trade_event_alert(event_type, stock, price, details=""):
    now_str = datetime.datetime.now().strftime("%H:%M:%S")
    if event_type == "ENTRY":
        msg = f"🚀 <b>ENTRY TRIGGERED</b> [{now_str}]\n\n<b>Stock:</b> {stock}\n<b>Fill Price:</b> ₹{price:.2f}\n<i>{details}</i>"
    elif event_type == "T1_PARTIAL":
        msg = f"🎯 <b>TARGET 1 HIT — 50% BOOKED</b> [{now_str}]\n\n<b>Stock:</b> {stock}\n<b>Booked 50% @:</b> ₹{price:.2f}\n<b>Breakeven SL Moved to:</b> {details}\n<i>Remaining 50% trailing to Target 2 risk-free!</i>"
    elif event_type == "TARGET_2":
        msg = f"🏆 <b>TARGET 2 HIT — 100% CLOSED</b> [{now_str}]\n\n<b>Stock:</b> {stock}\n<b>Exit Price:</b> ₹{price:.2f}\n<i>{details}</i>"
    elif event_type == "SL_EXIT":
        msg = f"🛑 <b>STOP LOSS TRIGGERED</b> [{now_str}]\n\n<b>Stock:</b> {stock}\n<b>Exit Fill:</b> ₹{price:.2f}\n<i>{details}</i>"
    elif event_type == "MANUAL":
        msg = f"⚡ <b>MANUAL POSITION EXIT</b> [{now_str}]\n\n<b>Stock:</b> {stock}\n<b>Exit Price:</b> ₹{price:.2f}\n<i>{details}</i>"
    else:
        msg = f"🔔 <b>Trade Alert [{stock}]:</b> {event_type} @ ₹{price:.2f} ({details})"
    send_telegram_message(msg)

def analyze_stock_for_telegram(symbol):
    sym = symbol.strip().upper()
    ticker_sym = sym if sym.endswith('.NS') else sym + '.NS'
    try:
        ticker = yf.Ticker(ticker_sym)
        hist = ticker.history(period="1y")
        if hist.empty:
            return f"❌ Could not find data for ticker <b>{sym}</b>."
        last_close = hist['Close'].iloc[-1]
        high_52    = hist['High'].max()
        low_52     = hist['Low'].min()
        pct_from_high = ((high_52 - last_close) / high_52) * 100
        sma50 = hist['Close'].rolling(50).mean().iloc[-1]
        sma50_5d_ago = hist['Close'].rolling(50).mean().iloc[-6] if len(hist) >= 56 else sma50
        slope_up = sma50 > sma50_5d_ago
        ret1m = ((last_close - hist['Close'].iloc[-21]) / hist['Close'].iloc[-21] * 100) if len(hist) >= 21 else 0.0
        delta = hist['Close'].diff()
        gain = delta.where(delta > 0, 0).rolling(14).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(14).mean()
        rs = gain / loss
        rsi = 100 - (100 / (1 + rs)).iloc[-1] if not rs.empty else 50.0

        if pct_from_high <= 0.5 and last_close > sma50 and slope_up:
            signal_badge = "🟢 <b>CONFIRMED BREAKOUT READY</b>"
        elif pct_from_high <= 3.0 and last_close > sma50:
            signal_badge = "🟡 <b>WATCHLIST NEAR-BREAKOUT (Coiling)</b>"
        elif last_close > sma50:
            signal_badge = "🔵 <b>BULLISH TRENDING (Above 50-DMA)</b>"
        else:
            signal_badge = "🔴 <b>BEARISH / BELOW 50-DMA</b>"

        news_item = ""
        try:
            news_list = ticker.news
            if news_list and len(news_list) > 0:
                top_news = news_list[0]
                title = top_news.get('title', top_news.get('headline', ''))
                publisher = top_news.get('publisher', top_news.get('source', 'News'))
                news_item = f"\n📰 <b>Latest News:</b> {title} (<i>{publisher}</i>)"
        except Exception: pass

        return (
            f"🔍 <b>ANALYSIS REPORT: {ticker_sym}</b>\nSignal: {signal_badge}\n───────────────────────────\n"
            f"• <b>LTP:</b> ₹{last_close:,.2f}\n• <b>52W High:</b> ₹{high_52:,.2f} ({pct_from_high:.1f}% below high)\n"
            f"• <b>52W Low:</b> ₹{low_52:,.2f}\n• <b>50-DMA:</b> ₹{sma50:,.2f} ({'📈 Rising' if slope_up else '📉 Falling'})\n"
            f"• <b>1-Month Return:</b> {ret1m:+.2f}%\n• <b>RSI (14):</b> {rsi:.1f}\n{news_item}"
        )
    except Exception as e:
        return f"❌ Error analyzing {sym}: {e}"

def generate_detailed_status_and_menu():
    if not os.path.exists(TRACKER_FILE):
        return "❌ Tracker file not found.", None
    try:
        wb = openpyxl.load_workbook(TRACKER_FILE, data_only=True)
        base_capital = 500000.0
        if 'Config' in wb.sheetnames:
            ws_c = wb['Config']
            for row in ws_c.iter_rows(min_row=2, max_row=15, values_only=True):
                if row[0] and ("total capital" in str(row[0]).lower() or str(row[0]).lower() == "capital"):
                    try: base_capital = float(str(row[1]).replace('%','').replace(',',''))
                    except Exception: pass
        df = pd.read_excel(TRACKER_FILE, sheet_name='Trade Log')
        if df.empty: return "ℹ️ No trade data found in tracker.", None
        open_only = df[df['Status'].astype(str).str.contains('OPEN', case=False, na=False)].copy()
        closed_only = df[df['Status'].astype(str).str.contains('CLOSED', case=False, na=False)].copy()
        total_pnl = closed_only['P&L (₹)'].sum() if not closed_only.empty else 0.0
        net_capital = base_capital + total_pnl
        win_rate = (len(closed_only[closed_only['P&L (₹)'] > 0]) / len(closed_only) * 100) if not closed_only.empty else 0.0
        deployed_cap = 0.0
        if not open_only.empty:
            for idx, r in open_only.iterrows():
                try:
                    qty = float(r['Qty']) if not pd.isna(r['Qty']) else 0.0
                    entry = float(r['Entry Price (₹)']) if not pd.isna(r['Entry Price (₹)']) else 0.0
                    deployed_cap += (qty * entry)
                except Exception: pass
        avail_capital = base_capital - deployed_cap + total_pnl
        open_stocks = open_only['Stock'].dropna().unique().tolist() if not open_only.empty else []
        live_prices = {}
        if open_stocks:
            tickers = [s + ".NS" if not s.endswith(".NS") else s for s in open_stocks]
            try:
                data = yf.download(tickers, period="1d", interval="1m", progress=False)
                if len(tickers) == 1:
                    sym = tickers[0]
                    if not data.empty: live_prices[sym] = float(data['Close'].iloc[-1])
                else:
                    for sym in tickers:
                        if sym in data['Close'] and not data['Close'][sym].dropna().empty:
                            live_prices[sym] = float(data['Close'][sym].dropna().iloc[-1])
            except Exception as e: print("Error fetching live prices:", e)

        msg = f"📊 <b>V2 INSTITUTIONAL PORTFOLIO STATUS REPORT</b>\n───────────────────────────\n"
        msg += f"💰 <b>Net Capital:</b> ₹{net_capital:,.2f}\n📈 <b>Realized P&L:</b> ₹{total_pnl:+,.2f} (Win Rate: {win_rate:.1f}%)\n"
        msg += f"💼 <b>Deployed Capital:</b> ₹{deployed_cap:,.2f}\n💵 <b>Available Buying Power:</b> ₹{avail_capital:,.2f}\n"
        tot_unrealized = 0.0
        position_details = ""
        inline_buttons = []
        if not open_only.empty:
            position_details += f"\n🟢 <b>ACTIVE OPEN POSITIONS ({len(open_only)}):</b>\n\n"
            for idx, r in open_only.reset_index().iterrows():
                sym = str(r['Stock'])
                ticker_sym = sym + ".NS" if not sym.endswith(".NS") else sym
                entry = float(r['Entry Price (₹)']) if not pd.isna(r['Entry Price (₹)']) else 0.0
                qty = float(r['Qty']) if not pd.isna(r['Qty']) else 0.0
                sl = float(r['SL Price (₹)']) if not pd.isna(r['SL Price (₹)']) else 0.0
                t1 = float(r['Target 1 (₹)']) if not pd.isna(r['Target 1 (₹)']) else 0.0
                t2 = float(r['Target 2 (₹)']) if not pd.isna(r['Target 2 (₹)']) else 0.0
                pos_size = float(r['Position Size (₹)']) if not pd.isna(r['Position Size (₹)']) else (qty * entry)
                sector = str(r['Sector']) if ('Sector' in r and not pd.isna(r['Sector'])) else "General"
                status_str = str(r['Status'])
                ltp = live_prices.get(ticker_sym, entry)
                unrealized = (ltp - entry) * qty - 20.0
                unrealized_pct = (unrealized / pos_size * 100) if pos_size > 0 else 0.0
                tot_unrealized += unrealized
                pnl_color = "🟢" if unrealized > 0 else "🔴"
                position_details += (
                    f"<b>{idx+1}. {sym}</b> ({sector}) [{status_str}]\n"
                    f"   • Bought At: ₹{entry:.2f} | LTP: ₹{ltp:.2f}\n"
                    f"   • Qty: {int(qty)} | Size: ₹{pos_size:,.2f}\n"
                    f"   • Live Net P&L: {pnl_color} <b>₹{unrealized:+,.2f} ({unrealized_pct:+.2f}%)</b>\n"
                    f"   • SL: ₹{sl:.2f} | T1: ₹{t1:.2f} | T2: ₹{t2:.2f}\n\n"
                )
                inline_buttons.append([{"text": f"🔴 Square Off {sym} (LTP ₹{ltp:.2f})", "callback_data": f"exit:{sym}"}])
            msg += f"⚡ <b>Live Net Unrealized P&L:</b> ₹{tot_unrealized:+,.2f}\n" + position_details
            msg += "👇 <b>Tap a button below to Square Off any trade:</b>"
        else:
            msg += "\nℹ️ <i>No active open positions.</i>"
        return msg, ({"inline_keyboard": inline_buttons} if inline_buttons else None)
    except Exception as e:
        return f"❌ Error generating status: {e}", None

def square_off_trade_from_telegram(sym):
    sym_clean = sym.strip().upper()
    ticker_sym = sym_clean + ".NS" if not sym_clean.endswith(".NS") else sym_clean
    exit_price = 0.0
    try:
        data = yf.download(ticker_sym, period="1d", interval="1m", progress=False)
        if not data.empty: exit_price = float(data['Close'].iloc[-1])
    except Exception: pass
    try:
        wb = openpyxl.load_workbook(TRACKER_FILE)
        ws = wb['Trade Log']
        found = False
        old_entry = 0.0
        for r in range(2, ws.max_row + 1):
            stk_val = ws.cell(row=r, column=2).value
            st_val = ws.cell(row=r, column=18).value
            if str(stk_val) == sym_clean and (st_val is None or "OPEN" in str(st_val)):
                old_entry = float(ws.cell(row=r, column=3).value or 0.0)
                if exit_price == 0.0: exit_price = float(ws.cell(row=r, column=19).value or old_entry)
                ws.cell(row=r, column=18).value = "CLOSED - MANUAL (Telegram)"
                ws.cell(row=r, column=19).value = exit_price
                found = True
                break
        if found:
            wb.save(TRACKER_FILE)
            return True, f"✅ <b>Successfully Squared Off {sym_clean}!</b>\n• Exit Price: ₹{exit_price:.2f}\n• Status: CLOSED - MANUAL (Telegram)"
        else:
            return False, f"⚠️ Could not find an active open trade for {sym_clean} in Excel."
    except Exception as err:
        return False, f"❌ Error executing square off: {err}"

def run_telegram_listener():
    bot_token, chat_id = get_telegram_config()
    if not bot_token: return
    print(f"[TELEGRAM BOT] Embedded listener online for Chat ID: {chat_id if chat_id else 'Any Chat'}...")
    try: requests.get(f"https://api.telegram.org/bot{bot_token}/deleteWebhook?drop_pending_updates=False", timeout=5)
    except Exception: pass

    offset = 0
    while True:
        try:
            url = f"https://api.telegram.org/bot{bot_token}/getUpdates?offset={offset}&timeout=10"
            r = requests.get(url, timeout=15)
            if r.status_code == 409:
                time.sleep(10)
                continue
            if r.status_code == 200:
                data = r.json()
                if "result" in data:
                    for update in data["result"]:
                        offset = update["update_id"] + 1
                        callback_query = update.get("callback_query", {})
                        if callback_query:
                            cb_id = callback_query.get("id")
                            cb_data = callback_query.get("data", "")
                            if cb_data.startswith("exit:"):
                                target_sym = cb_data.split(":", 1)[1]
                                answer_callback_query(cb_id, text=f"Squaring off {target_sym}...")
                                _, result_msg = square_off_trade_from_telegram(target_sym)
                                send_telegram_message(result_msg)
                                status_msg, status_markup = generate_detailed_status_and_menu()
                                if status_markup: send_telegram_message_with_keyboard(status_msg, status_markup)
                                else: send_telegram_message(status_msg)
                                continue

                        message = update.get("message", update.get("channel_post", {}))
                        text = message.get("text", "").strip()
                        msg_chat_id = str(message.get("chat", {}).get("id", ""))
                        if not text: continue
                        print(f"-> [TELEGRAM RECV] Chat {msg_chat_id}: '{text}'")
                        cmd = text.lower().strip()

                        if cmd in ["/start", "/help"]:
                            send_telegram_message(
                                "🤖 <b>V2 Trading Bot Active!</b>\n\n"
                                "You can message me <b>any stock name</b> (e.g. <code>RELIANCE</code>, <code>TATAMOTORS</code>) to analyze it instantly!\n\n"
                                "Or use commands:\n"
                                "• <code>/status</code> or <code>/squareoff</code> — Detailed P&L report & Square Off menu\n"
                                "• <code>/picks</code> — Tomorrow's allocated trades\n"
                                "• <code>/watchlist</code> — Near-breakout watchlist"
                            )
                        elif cmd == "/picks":
                            if os.path.exists(TRACKER_FILE):
                                df = pd.read_excel(TRACKER_FILE, sheet_name='Trade Log')
                                pending = df[df['Status'].isna() | (df['Status'] == 'PENDING')]
                                if not pending.empty:
                                    reply = "📋 <b>Pending Trades Allocated for Tomorrow:</b>\n\n"
                                    for idx, r_row in pending.iterrows():
                                        reply += f"• <b>{r_row['Stock']}</b> | Entry: ₹{r_row['Entry Price (₹)']} | SL: ₹{r_row['SL Price (₹)']} | T2: ₹{r_row['Target 2 (₹)']}\n"
                                else: reply = "ℹ️ No pending trades allocated for tomorrow."
                            else: reply = "❌ Tracker file not found."
                            send_telegram_message(reply)

                        elif cmd == "/watchlist":
                            if os.path.exists(TRACKER_FILE):
                                xl = pd.ExcelFile(TRACKER_FILE)
                                if 'Watchlist' in xl.sheet_names:
                                    wl_df = pd.read_excel(TRACKER_FILE, sheet_name='Watchlist')
                                    if not wl_df.empty and 'Stock' in wl_df.columns:
                                        reply = "👀 <b>Top Near-Breakout Watchlist:</b>\n\n"
                                        for idx, r_row in wl_df.head(8).iterrows():
                                            reply += f"• <b>{r_row['Stock']}</b>: ₹{r_row['LTP (₹)']} ({r_row['% from High']:.1f}% from 52W High)\n"
                                    else: reply = "ℹ️ Watchlist is currently empty."
                                else: reply = "ℹ️ Watchlist sheet not found."
                            else: reply = "❌ Tracker file not found."
                            send_telegram_message(reply)

                        elif cmd in ["/status", "/squareoff", "/exit"]:
                            status_msg, status_markup = generate_detailed_status_and_menu()
                            if status_markup: send_telegram_message_with_keyboard(status_msg, status_markup)
                            else: send_telegram_message(status_msg)

                        else:
                            clean_target = cmd.replace("/analyze", "").strip().upper()
                            if clean_target:
                                send_telegram_message(f"🔍 Analyzing <b>{clean_target}</b>...")
                                reply = analyze_stock_for_telegram(clean_target)
                                send_telegram_message(reply)
            time.sleep(2)
        except Exception as e:
            print("Error in Telegram listener loop:", e)
            time.sleep(5)

@st.cache_resource
def start_background_paper_trader():
    def trader_loop():
        while True:
            try:
                run_paper_trader()
            except Exception as e:
                print("Error in background trader:", e)
            time.sleep(POLL_INTERVAL)
            
    t = threading.Thread(target=trader_loop, daemon=True)
    t.start()
    return t

@st.cache_resource
def start_unified_telegram_listener():
    if getattr(sys, '_v2_telegram_listener_started', False):
        return None
    sys._v2_telegram_listener_started = True
    try:
        t = threading.Thread(target=run_telegram_listener, daemon=True)
        t.start()
        print("[DASHBOARD] Unified Telegram Bot command listener started in background thread.")
        return t
    except Exception as e:
        print("[DASHBOARD] Error starting Telegram listener thread:", e)
        return None

@st.cache_resource
def start_auto_daily_screener_scheduler():
    if getattr(sys, '_v2_screener_scheduler_started', False):
        return None
    sys._v2_screener_scheduler_started = True

    def screener_schedule_loop():
        print("[AUTO SCREENER] Daily 4:00 PM Screener Scheduler Active.")
        while True:
            now = datetime.datetime.now()
            if now.hour == 16 and now.minute == 0 and now.weekday() < 5:
                print("[AUTO SCREENER] 4:00 PM reached! Running v2_engine.py auto-scan...")
                try:
                    import subprocess
                    subprocess.run([sys.executable, os.path.join(V2_DIR, "v2_engine.py")], check=True)
                    print("[AUTO SCREENER] Auto-scan completed & Telegram alert sent!")
                except Exception as e:
                    print("[AUTO SCREENER ERROR]", e)
                time.sleep(60)
            time.sleep(30)
            
    try:
        t = threading.Thread(target=screener_schedule_loop, daemon=True)
        t.start()
        print("[DASHBOARD] Auto 4:00 PM Screener Scheduler started in background thread.")
        return t
    except Exception as e:
        print("[DASHBOARD] Error starting screener scheduler:", e)
        return None

# Initialize background services
start_background_paper_trader()
start_unified_telegram_listener()
start_auto_daily_screener_scheduler()

# ==========================================
# 3. DASHBOARD UI
# ==========================================
@st.cache_data(ttl=5)
def load_data():
    if not os.path.exists(TRACKER_FILE):
        return pd.DataFrame()
    try:
        return pd.read_excel(TRACKER_FILE, sheet_name='Trade Log')
    except Exception as e:
        st.error(f"Error reading tracker: {e}")
        return pd.DataFrame()

@st.cache_data(ttl=5)
def load_watchlist():
    if not os.path.exists(TRACKER_FILE):
        return pd.DataFrame()
    try:
        xl = pd.ExcelFile(TRACKER_FILE)
        if 'Watchlist' in xl.sheet_names:
            return pd.read_excel(TRACKER_FILE, sheet_name='Watchlist')
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()

# Sidebar Configuration
with st.sidebar:
    st.image("https://cdn-icons-png.flaticon.com/512/2422/2422796.png", width=100)
    st.markdown("## Configuration")
    base_capital = st.number_input("Base Capital (₹)", min_value=10000.0, value=500000.0, step=10000.0)
    
    st.markdown("---")
    st.markdown("### Date Filter")
    date_filter = st.selectbox("View History", ["All Time", "Last 30 Days", "Last 90 Days", "This Year"])
    
    if st.button("Refresh Data 🔄", width="stretch"):
        st.cache_data.clear()

st.title("V2 Paper Trading Dashboard 🚀")
st.caption("Live trading engine is running in the background.")

df = load_data()
if df.empty:
    st.warning("No data found in tracker.xlsx.")
    st.stop()

df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
if date_filter == "Last 30 Days": df = df[df['Date'] >= pd.Timestamp.now() - pd.Timedelta(days=30)]
elif date_filter == "Last 90 Days": df = df[df['Date'] >= pd.Timestamp.now() - pd.Timedelta(days=90)]
elif date_filter == "This Year": df = df[df['Date'] >= pd.Timestamp(pd.Timestamp.now().year, 1, 1)]

open_trades = df[df['Status'].astype(str).str.contains('OPEN|PENDING', case=False, regex=True) | df['Status'].isna()]
closed_trades = df[df['Status'].astype(str).str.contains('CLOSED', case=False, na=False)].copy()

total_pnl = closed_trades['P&L (₹)'].sum() if not closed_trades.empty else 0.0
net_capital = base_capital + total_pnl

if not closed_trades.empty:
    winning_trades = closed_trades[closed_trades['P&L (₹)'] > 0]
    losing_trades = closed_trades[closed_trades['P&L (₹)'] <= 0]
    win_rate = (len(winning_trades) / len(closed_trades)) * 100
    
    gross_profit, gross_loss = winning_trades['P&L (₹)'].sum(), abs(losing_trades['P&L (₹)'].sum())
    profit_factor = (gross_profit / gross_loss) if gross_loss != 0 else float('inf')
    
    returns = closed_trades['P&L (₹)'] / base_capital
    sharpe_ratio = (returns.mean() / returns.std()) * np.sqrt(100) if len(returns) > 1 and returns.std() != 0 else 0.0
    
    # ── Institutional Metrics ──
    ct_sorted = closed_trades.sort_values(by='Date').copy()
    ct_sorted['Equity'] = base_capital + ct_sorted['P&L (₹)'].cumsum()
    peak = ct_sorted['Equity'].cummax()
    dd_pct = (ct_sorted['Equity'] - peak) / peak * 100
    max_drawdown = abs(dd_pct.min()) if not dd_pct.empty else 0.0
    
    tot_return_pct = (total_pnl / base_capital) * 100
    calmar_ratio = (tot_return_pct / max_drawdown) if max_drawdown > 0 else (tot_return_pct if tot_return_pct > 0 else 0.0)
    
    avg_win_r = winning_trades['R:R (T1)'].mean() if ('R:R (T1)' in winning_trades and not winning_trades.empty and not pd.isna(winning_trades['R:R (T1)'].mean())) else 1.0
    win_dec = win_rate / 100.0
    loss_dec = (100.0 - win_rate) / 100.0
    expectancy_r = (win_dec * avg_win_r) - (loss_dec * 1.0)
else:
    win_rate, profit_factor, sharpe_ratio = 0.0, 0.0, 0.0
    max_drawdown, calmar_ratio, expectancy_r = 0.0, 0.0, 0.0

col1, col2, col3, col4, col5 = st.columns(5)
col1.metric("Net Capital (₹)", f"₹{net_capital:,.2f}", f"{total_pnl:,.2f}")
col2.metric("Total P&L", f"₹{total_pnl:,.2f}", f"{(total_pnl/base_capital)*100:.1f}%")
col3.metric("Win Rate", f"{win_rate:.1f}%")
col4.metric("Profit Factor", f"{profit_factor:.2f}")
col5.metric("Est. Sharpe Ratio", f"{sharpe_ratio:.2f}")

st.markdown("<br>", unsafe_allow_html=True)
inst1, inst2, inst3 = st.columns(3)
inst1.metric("Max Drawdown (MDD)", f"{max_drawdown:.2f}%")
inst2.metric("Calmar Ratio", f"{calmar_ratio:.2f}")
inst3.metric("Expectancy (R-Multiple)", f"{expectancy_r:+.2f} R / trade")

st.markdown("---")
tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Overview & Equity Curve", "🟢 Current Trades", "🔴 Closed Trades", "📈 Stock Analysis", "👀 Watchlist"])

with tab1:
    st.markdown("### Portfolio Performance")
    if not closed_trades.empty:
        ct = closed_trades.sort_values(by='Date').copy()
        ct['Cumulative P&L'] = ct['P&L (₹)'].cumsum()
        
        fig = px.area(ct, x='Date', y='Cumulative P&L', title="Cumulative Equity Curve", color_discrete_sequence=['#00F0FF'])
        fig.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', font_color='#FFF', hovermode="x unified")
        fig.update_traces(fill='tozeroy', fillcolor='rgba(0, 240, 255, 0.2)', line=dict(width=3))
        st.plotly_chart(fig, width="stretch")
        
        c1, c2 = st.columns(2)
        with c1:
            fig_pie = px.pie(values=[len(winning_trades), len(losing_trades)], names=['Wins', 'Losses'], title="Win / Loss", color_discrete_sequence=['#00FF88', '#FF3366'], hole=0.4)
            fig_pie.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', font_color='#FFF')
            st.plotly_chart(fig_pie, width="stretch")
        with c2:
            fig_hist = px.histogram(closed_trades, x='P&L (₹)', nbins=20, title="P&L Distribution", color_discrete_sequence=['#00F0FF'])
            fig_hist.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', font_color='#FFF')
            st.plotly_chart(fig_hist, width="stretch")
    else:
        st.info("No closed trades yet.")

with tab2:
    st.markdown("### Active & Pending Trades")
    if not open_trades.empty:
        unique_stocks = [str(s) for s in open_trades['Stock'].dropna().unique()]
        tickers = [s + ".NS" if not s.endswith(".NS") else s for s in unique_stocks]
        live_prices = fetch_live_prices(tickers)
        
        display_rows = []
        tot_unrealized_pnl = 0.0
        tot_deployed = 0.0
        active_open_count = 0
        
        for idx, row in open_trades.iterrows():
            sym = str(row['Stock'])
            ticker_sym = sym + ".NS" if not sym.endswith(".NS") else sym
            status = str(row['Status']) if not pd.isna(row['Status']) else "PENDING"
            
            entry = float(row['Entry Price (₹)']) if not pd.isna(row['Entry Price (₹)']) else 0.0
            qty = float(row['Qty']) if not pd.isna(row['Qty']) else 0.0
            pos_size = float(row['Position Size (₹)']) if not pd.isna(row['Position Size (₹)']) else 0.0
            
            ltp = entry
            unrealized_pnl = None
            unrealized_pnl_pct = None
            
            if ticker_sym in live_prices and 'LTP' in live_prices[ticker_sym]:
                ltp = round(float(live_prices[ticker_sym]['LTP']), 2)
                if status == "OPEN" and entry > 0 and qty > 0:
                    # Deduct ₹20 buy order brokerage already incurred on entry
                    unrealized_pnl = round((ltp - entry) * qty - 20.0, 2)
                    unrealized_pnl_pct = round((unrealized_pnl / pos_size) * 100, 2) if pos_size > 0 else 0.0
                    tot_unrealized_pnl += unrealized_pnl
                    tot_deployed += pos_size
                    active_open_count += 1
            
            date_str = pd.to_datetime(row['Date']).strftime('%Y-%m-%d') if not pd.isna(row['Date']) else ""
            clean_sym = sym[:-3] if sym.endswith('.NS') else sym
            yf_link = f"https://finance.yahoo.com/quote/{clean_sym}.NS"
            
            display_rows.append({
                'Date': date_str,
                'Stock': sym,
                'Status': status,
                'Bought At (₹)': entry,
                'Current Price (₹)': ltp,
                'Unrealized P&L (₹)': unrealized_pnl if unrealized_pnl is not None else np.nan,
                'Unrealized P&L %': unrealized_pnl_pct if unrealized_pnl_pct is not None else np.nan,
                'SL Price (₹)': row['SL Price (₹)'],
                'Target 1 (₹)': row['Target 1 (₹)'],
                'Target 2 (₹)': row['Target 2 (₹)'],
                'Qty': qty,
                'Position Size (₹)': pos_size,
                'Risk Amt (₹)': row['Risk Amt (₹)'],
                'Yahoo Finance': yf_link
            })
            
        display_df = pd.DataFrame(display_rows)
        # Ensure all numeric columns are strictly float (Arrow-compatible)
        num_cols = ['Bought At (₹)', 'Current Price (₹)', 'Unrealized P&L (₹)', 'Unrealized P&L %', 'SL Price (₹)', 'Target 1 (₹)', 'Target 2 (₹)', 'Qty', 'Position Size (₹)', 'Risk Amt (₹)']
        for col in num_cols:
            if col in display_df.columns:
                display_df[col] = pd.to_numeric(display_df[col], errors='coerce')
        
        op1, op2, op3 = st.columns(3)
        op1.metric("Active Open Positions", f"{active_open_count}")
        op2.metric("Deployed Capital", f"₹{tot_deployed:,.2f}")
        op3.metric("Live Net P&L (after Buy Fee)", f"₹{tot_unrealized_pnl:,.2f}", f"{(tot_unrealized_pnl / tot_deployed * 100) if tot_deployed > 0 else 0:.2f}%")
        
        st.markdown("---")
        
        def highlight_pnl_val(val):
            if isinstance(val, (float, int)) and not pd.isna(val):
                return f"color: {'#00FF88' if val > 0 else '#FF3366'}; font-weight: bold;"
            return ''
            
        st.dataframe(
            display_df.style.map(highlight_pnl_val, subset=['Unrealized P&L (₹)', 'Unrealized P&L %']),
            column_config={
                "Yahoo Finance": st.column_config.LinkColumn("Yahoo Finance", display_text="🔗 View Graph & Data"),
                "Unrealized P&L (₹)": st.column_config.NumberColumn("Unrealized P&L (₹)", format="₹%.2f"),
                "Unrealized P&L %": st.column_config.NumberColumn("Unrealized P&L %", format="%.2f%%")
            },
            width="stretch",
            hide_index=True
        )
        
        # ⚡ Manual Exit Control Panel
        open_only = open_trades[open_trades['Status'] == 'OPEN']
        if not open_only.empty:
            st.markdown("---")
            st.markdown("### ⚡ Manual Exit Control Panel")
            st.caption("Need to exit an active position early? Select the stock below to manually close it right now at market or limit price.")
            
            ex_c1, ex_c2, ex_c3, ex_c4 = st.columns([2, 1.5, 1.5, 1])
            with ex_c1:
                selected_exit_stock = st.selectbox("Select Active Stock to Exit", open_only['Stock'].dropna().unique(), key="exit_stock_select")
                
            selected_row = open_only[open_only['Stock'] == selected_exit_stock].iloc[0]
            stk_name = str(selected_exit_stock)
            t_name = stk_name + ".NS" if not stk_name.endswith(".NS") else stk_name
            default_exit_price = live_prices.get(t_name, {}).get('LTP', float(selected_row['Entry Price (₹)']))
            
            with ex_c2:
                user_exit_price = st.number_input("Exit Price (₹)", value=float(default_exit_price), step=0.5, key="exit_price_input")
            with ex_c3:
                user_exit_reason = st.selectbox("Exit Reason", ["CLOSED - MANUAL", "CLOSED - SL", "CLOSED - TARGET 1", "CLOSED - TARGET 2"], key="exit_reason_select")
            with ex_c4:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("Confirm Exit 🔴", key="confirm_manual_exit_btn", width="stretch"):
                    try:
                        wb_mod = openpyxl.load_workbook(TRACKER_FILE)
                        ws_mod = wb_mod['Trade Log']
                        found_row = False
                        for r in range(2, ws_mod.max_row + 1):
                            s_val = ws_mod.cell(row=r, column=2).value
                            st_val = ws_mod.cell(row=r, column=18).value
                            if str(s_val) == stk_name and (st_val is None or st_val == "OPEN"):
                                ws_mod.cell(row=r, column=18).value = user_exit_reason
                                ws_mod.cell(row=r, column=19).value = float(user_exit_price)
                                found_row = True
                                break
                        if found_row:
                            wb_mod.save(TRACKER_FILE)
                            st.success(f"Successfully closed {stk_name} at ₹{user_exit_price:.2f} ({user_exit_reason})!")
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.error(f"Could not locate active row for {stk_name} in Excel.")
                    except Exception as err:
                        st.error(f"Error closing position: {err}")
    else:
        st.info("No active trades found.")

with tab3:
    st.markdown("### Trade History")
    if not closed_trades.empty:
        display_df = closed_trades[['Date', 'Stock', 'Status', 'Entry Price (₹)', 'Exit Price (₹)', 'P&L (₹)', 'P&L %']].copy()
        display_df['Date'] = display_df['Date'].dt.strftime('%Y-%m-%d')
        display_df['Yahoo Finance'] = display_df['Stock'].apply(lambda s: f"https://finance.yahoo.com/quote/{s if not str(s).endswith('.NS') else str(s)[:-3]}.NS")
        
        def highlight_pnl(val):
            if type(val) in [float, int] and not pd.isna(val):
                return f"color: {'#00FF88' if val > 0 else '#FF3366'}; font-weight: bold;"
            return ''
            
        st.dataframe(
            display_df.style.map(highlight_pnl, subset=['P&L (₹)', 'P&L %']),
            column_config={
                "Yahoo Finance": st.column_config.LinkColumn("Yahoo Finance", display_text="🔗 View Graph & Data")
            },
            width="stretch",
            hide_index=True
        )
    else:
        st.info("No closed trades found.")

with tab4:
    st.markdown("### 📈 Professional Stock Analysis & Signals")
    
    all_stocks = df['Stock'].dropna().unique().tolist() if not df.empty else []
    
    scol1, scol2 = st.columns([1, 2])
    with scol1:
        selected_stock = st.selectbox("Select a stock from your portfolio", ["-- Select --"] + all_stocks)
    with scol2:
        custom_stock = st.text_input("Or search any NSE Ticker (e.g. RELIANCE)")
        
    target_stock = custom_stock.strip().upper() if custom_stock.strip() else selected_stock
    
    if target_stock and target_stock != "-- Select --":
        sym = target_stock if target_stock.endswith('.NS') else target_stock + '.NS'
        raw_ticker = sym[:-3] if sym.endswith('.NS') else sym
        st.markdown(f"**Analyzing:** `{sym}`")
        
        with st.spinner(f"Analyzing {sym} and fetching live market news..."):
            try:
                ticker = yf.Ticker(sym)
                hist = ticker.history(period="1y")
                
                if not hist.empty:
                    last_price = hist['Close'].iloc[-1]
                    high_52 = hist['High'].max()
                    low_52 = hist['Low'].min()
                    avg_vol = hist['Volume'].mean()
                    pct_from_high = ((high_52 - last_price) / high_52) * 100
                    
                    hist['50_MA'] = hist['Close'].rolling(window=50).mean()
                    sma50_last = hist['50_MA'].iloc[-1]
                    sma50_5d_ago = hist['50_MA'].iloc[-6] if len(hist) >= 56 else sma50_last
                    slope_up = sma50_last > sma50_5d_ago
                    
                    ret1m = ((last_price - hist['Close'].iloc[-21]) / hist['Close'].iloc[-21] * 100) if len(hist) >= 21 else 0.0
                    
                    # RSI 14
                    delta = hist['Close'].diff()
                    gain = delta.where(delta > 0, 0).rolling(14).mean()
                    loss = (-delta.where(delta < 0, 0)).rolling(14).mean()
                    rs = gain / loss
                    rsi_14 = 100 - (100 / (1 + rs)).iloc[-1] if not rs.empty else 50.0
                    
                    # Strategy Signal Card
                    if pct_from_high <= 0.5 and last_price > sma50_last and slope_up:
                        sig_label = "🟢 QUALIFIED CONFIRMED BREAKOUT"
                        sig_desc = "Stock is breaking out at 52W High with 50-DMA upward slope & momentum alignment."
                        sig_color = "rgba(0, 255, 136, 0.15)"
                        border_color = "#00FF88"
                    elif pct_from_high <= 3.0 and last_price > sma50_last:
                        sig_label = "🟡 WATCHLIST NEAR-BREAKOUT (Coiling)"
                        sig_desc = "Stock is coiling within 3% of 52W High. Watch for volume surge trigger."
                        sig_color = "rgba(255, 215, 0, 0.15)"
                        border_color = "#FFD700"
                    elif last_price > sma50_last:
                        sig_label = "🔵 BULLISH TRENDING"
                        sig_desc = "Trading above 50-DMA, but not in immediate 52W High breakout zone."
                        sig_color = "rgba(0, 240, 255, 0.15)"
                        border_color = "#00F0FF"
                    else:
                        sig_label = "🔴 BEARISH / BELOW 50-DMA"
                        sig_desc = "Trading below 50-DMA. Fails system momentum trend filters."
                        sig_color = "rgba(255, 51, 102, 0.15)"
                        border_color = "#FF3366"
                        
                    st.markdown(f"""
                    <div style="background:{sig_color}; border:2px solid {border_color}; padding:15px; border-radius:12px; margin-bottom:15px;">
                        <h3 style="margin:0; color:#FFF;">{sig_label}</h3>
                        <p style="margin:5px 0 0 0; color:#DDD;">{sig_desc}</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    mc1, mc2, mc3, mc4, mc5 = st.columns(5)
                    mc1.metric("LTP (₹)", f"₹{last_price:,.2f}")
                    mc2.metric("52W High", f"₹{high_52:,.2f}", f"{pct_from_high:.1f}% below")
                    mc3.metric("50-DMA", f"₹{sma50_last:,.2f}", "Rising" if slope_up else "Falling")
                    mc4.metric("1M Return", f"{ret1m:+.2f}%")
                    mc5.metric("RSI (14)", f"{rsi_14:.1f}")
                    
                    st.markdown("---")
                    
                    # Candlestick Chart
                    fig = go.Figure(data=[go.Candlestick(
                        x=hist.index,
                        open=hist['Open'], high=hist['High'],
                        low=hist['Low'], close=hist['Close'],
                        name="Price"
                    )])
                    fig.add_trace(go.Scatter(x=hist.index, y=hist['50_MA'], line=dict(color='#00F0FF', width=1.5), name='50 DMA'))
                    fig.update_layout(
                        title=f"{sym} - 1 Year Price Action & 50 DMA",
                        yaxis_title="Price (₹)", xaxis_title="Date",
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        font_color='#FFF', height=550, margin=dict(l=0, r=0, t=40, b=0),
                        xaxis_rangeslider_visible=False
                    )
                    st.plotly_chart(fig, width="stretch")
                    
                    # 📰 Market News Feed Section
                    st.markdown("---")
                    st.markdown(f"### 📰 Latest Market News & Headlines for {raw_ticker}")
                    try:
                        news_items = ticker.news
                        if news_items and len(news_items) > 0:
                            for item in news_items[:4]:
                                title = item.get('title', item.get('headline', 'Market Update'))
                                publisher = item.get('publisher', item.get('source', 'Financial News'))
                                link = item.get('link', f"https://finance.yahoo.com/quote/{sym}")
                                pub_time = item.get('providerPublishTime', None)
                                time_str = datetime.datetime.fromtimestamp(pub_time).strftime('%d %b %Y, %H:%M') if pub_time else ""
                                
                                st.markdown(f"""
                                <div style="background:rgba(255,255,255,0.05); border-left:4px solid #00F0FF; padding:12px 16px; margin-bottom:10px; border-radius:6px;">
                                    <h4 style="margin:0; font-size:1.1rem;"><a href="{link}" target="_blank" style="color:#00F0FF; text-decoration:none;">{title}</a></h4>
                                    <p style="margin:4px 0 0 0; color:#AAA; font-size:0.85rem;">Source: <b>{publisher}</b> {f'• {time_str}' if time_str else ''}</p>
                                </div>
                                """, unsafe_allow_html=True)
                        else:
                            st.info(f"No recent news articles found for {raw_ticker}. <a href='https://finance.yahoo.com/quote/{sym}' target='_blank'>View on Yahoo Finance</a>", unsafe_allow_html=True)
                    except Exception as n_err:
                        st.info(f"Could not load news feed: {n_err}")
                else:
                    st.error(f"No data found for {sym}. Please check ticker symbol.")
            except Exception as e:
                st.error(f"Error fetching data: {e}")
    else:
        st.info("👆 Select a stock from the dropdown or type a custom ticker to view analysis, strategy signals, and live news.")

with tab5:
    st.markdown("### 👀 Near-Breakout Watchlist")
    st.caption("These stocks are coiling within 3% of their 52-week high. They did NOT qualify for auto-trade today. Watch them manually — a volume surge + close above the high tomorrow = potential next day pick.")

    wl_df = load_watchlist()

    if not wl_df.empty and wl_df.shape[1] > 1:
        # Style helper: highlight % from high (closer to 0 = more imminent)
        def highlight_proximity(val):
            if isinstance(val, (float, int)):
                if val < 1.0:
                    return 'color: #FF8C00; font-weight: bold;'   # orange — very close!
                elif val < 2.0:
                    return 'color: #FFD700; font-weight: bold;'   # yellow
            return ''

        # Drop 'Note' for the main table — show it as caption below
        display_cols = [c for c in wl_df.columns if c != 'Note']
        display_df = wl_df[display_cols].copy() if display_cols else wl_df.copy()
        if 'Stock' in display_df.columns:
            display_df['Yahoo Finance'] = display_df['Stock'].apply(lambda s: f"https://finance.yahoo.com/quote/{s if not str(s).endswith('.NS') else str(s)[:-3]}.NS")

        styled = display_df.style.map(highlight_proximity, subset=['% from High']) \
            if '% from High' in display_df.columns else display_df.style

        st.dataframe(
            styled,
            column_config={
                "Yahoo Finance": st.column_config.LinkColumn("Yahoo Finance", display_text="🔗 View Graph & Data")
            },
            width="stretch",
            hide_index=True
        )

        # Bar chart: distance from 52W High
        if '% from High' in wl_df.columns and 'Stock' in wl_df.columns:
            fig_wl = px.bar(
                wl_df, x='Stock', y='% from High',
                title="Distance from 52-Week High (%)",
                color='% from High',
                color_continuous_scale=['#00FF88', '#FFD700', '#FF3366'],
                text='% from High'
            )
            fig_wl.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
            fig_wl.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font_color='#FFF', showlegend=False,
                yaxis_title="% Below 52W High", xaxis_title="Stock",
                coloraxis_showscale=False
            )
            st.plotly_chart(fig_wl, width="stretch")

        st.info("💡 Tip: Sort by '% from High' ascending to see the stocks closest to triggering. An orange row means < 1% away — check the chart manually!")
    else:
        st.info("No watchlist data found. Run `python v2_engine.py` to populate this.")
