import os
import sys
import datetime
import math
import requests
from bs4 import BeautifulSoup
import yfinance as yf
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Force utf-8 encoding for Windows terminals
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

V2_DIR = os.path.dirname(os.path.abspath(__file__))
QUERY_FILE           = os.path.join(V2_DIR, "chartink query.txt")
WATCHLIST_QUERY_FILE = os.path.join(V2_DIR, "chartink watchlist query.txt")
TRACKER_FILE         = os.path.join(V2_DIR, "tracker.xlsx")

# ─────────────────────────────────────────────
# CONFIGURATION (tunable thresholds)
# ─────────────────────────────────────────────
SCORE_THRESHOLD = 1.0   # Relative momentum vs Nifty (was 0.0 — too loose)
MAX_SL_PCT      = 7.0   # Reject trades where stop-loss > 7% away
MAX_TRADES      = 5     # Max confirmed breakout trades per day

# ─────────────────────────────────────────────
# CHARTINK SCRAPER
# ─────────────────────────────────────────────
def get_chartink_stocks(query, label="candidates"):
    print(f"-> Fetching {label} from Chartink...")
    clean_lines = []
    for line in query.split('\n'):
        clean_line = line.split('//')[0].strip()
        if clean_line:
            clean_lines.append(clean_line)
    clean_query = ' '.join(clean_lines)

    with requests.Session() as s:
        r = s.get('https://chartink.com/screener/time-pass')
        soup = BeautifulSoup(r.text, 'html.parser')
        csrf_meta = soup.select_one('meta[name="csrf-token"]')
        if not csrf_meta:
            print("Error: Could not find CSRF token on Chartink.")
            return []
        csrf = csrf_meta['content']
        url  = 'https://chartink.com/screener/process'
        headers  = {'X-CSRF-TOKEN': csrf, 'X-Requested-With': 'XMLHttpRequest'}
        response = s.post(url, data={'scan_clause': clean_query}, headers=headers)
        if response.status_code != 200:
            print(f"Chartink API returned {response.status_code}")
            return []
        resp_json = response.json()
        if 'data' in resp_json:
            stocks = [row['nsecode'] for row in resp_json['data']]
            print(f"-> Chartink found {len(stocks)} {label}: {stocks}")
            return stocks
        return []

# ─────────────────────────────────────────────
# BENCHMARK DATA (cached once per run)
# ─────────────────────────────────────────────
def get_benchmark():
    print("-> Fetching benchmark data (^NSEI)...")
    bench = yf.Ticker("^NSEI").history(period="3y")
    if bench.empty:
        print("Error fetching benchmark data.")
        return None, None, False

    bench['DailyRet'] = bench['Close'].pct_change() * 100
    bench['Ret6m']    = (bench['Close'] - bench['Close'].shift(126)) / bench['Close'].shift(126) * 100
    bench['Ret12m']   = (bench['Close'] - bench['Close'].shift(252)) / bench['Close'].shift(252) * 100
    bench['Vol6m']    = bench['DailyRet'].rolling(126).std()
    bench['Vol12m']   = bench['DailyRet'].rolling(252).std()
    bench['Score']    = (bench['Ret6m'] / bench['Vol6m']) + (bench['Ret12m'] / bench['Vol12m'])

    # Weekly regime: Nifty above 200-week SMA
    bench_weekly = yf.Ticker("^NSEI").history(period="8y", interval="1wk")
    regime_ok = True
    if not bench_weekly.empty:
        bench_weekly['SMA200'] = bench_weekly['Close'].rolling(200).mean()
        last_w = bench_weekly.iloc[-1]
        regime_ok = last_w['Close'] > last_w['SMA200']

    return bench.iloc[-1], bench, regime_ok

# ─────────────────────────────────────────────
# EVALUATE A STOCK FOR TRADE QUALIFICATION
# ─────────────────────────────────────────────
def evaluate_stock(sym, bench_last, bench_df, regime_ok):
    """
    Returns a trade dict if the stock passes all filters, else None.
    Filters applied (in order of rejection speed):
      1. Enough data (>252 bars)
      2. Market regime OK (Nifty weekly > 200 SMA)
      3. Relative momentum score > SCORE_THRESHOLD
      4. Price above SMA50 (trend filter)
      5. SMA50 is rising (slope filter — prevents catching dead-cat bounces)
      6. 1-month return > 0 (near-term momentum confirmation)
      7. SL% <= MAX_SL_PCT (position sizing quality gate)
    """
    ticker = sym + ".NS"
    data = yf.Ticker(ticker).history(period="3y")

    if len(data) < 252:
        print(f"  [SKIP] {sym}: not enough data ({len(data)} bars)")
        return None

    # ── Momentum score ──
    data['DailyRet'] = data['Close'].pct_change() * 100
    data['Ret6m']    = (data['Close'] - data['Close'].shift(126)) / data['Close'].shift(126) * 100
    data['Ret12m']   = (data['Close'] - data['Close'].shift(252)) / data['Close'].shift(252) * 100
    data['Vol6m']    = data['DailyRet'].rolling(126).std()
    data['Vol12m']   = data['DailyRet'].rolling(252).std()
    data['Score']    = (data['Ret6m'] / data['Vol6m']) + (data['Ret12m'] / data['Vol12m'])

    # ── Trend / slope ──
    data['SMA50']    = data['Close'].rolling(50).mean()
    data['SMA50_5d'] = data['SMA50'].shift(5)

    # ── ATR14 ──
    data['PrevClose'] = data['Close'].shift(1)
    data['TR'] = data.apply(
        lambda r: max(r['High'] - r['Low'],
                      abs(r['High'] - r['PrevClose']),
                      abs(r['Low']  - r['PrevClose']))
        if not pd.isna(r['PrevClose']) else r['High'] - r['Low'], axis=1)
    data['ATR14'] = data['TR'].ewm(alpha=1/14, adjust=False).mean()

    # ── 1-month return ──
    data['Ret1m'] = (data['Close'] - data['Close'].shift(21)) / data['Close'].shift(21) * 100

    last = data.iloc[-1]

    relative_score = last['Score'] - bench_last['Score']
    trend_ok       = last['Close'] > last['SMA50']
    slope_ok       = last['SMA50'] > data['SMA50_5d'].iloc[-1]   # SMA50 is rising
    momentum_ok    = data['Ret1m'].iloc[-1] > 0                   # 1-month positive

    reasons = []
    if not regime_ok:     reasons.append("regime")
    if relative_score <= SCORE_THRESHOLD: reasons.append(f"score={relative_score:.2f}<{SCORE_THRESHOLD}")
    if not trend_ok:      reasons.append("below_SMA50")
    if not slope_ok:      reasons.append("SMA50_falling")
    if not momentum_ok:   reasons.append("1m_neg")

    if reasons:
        print(f"  [REJECTED] {sym}: {', '.join(reasons)}")
        return None

    entry    = last['Close']
    atr_val  = last['ATR14']
    sl       = entry - (atr_val * 2.0)
    sl_width = entry - sl
    sl_pct   = (sl_width / entry) * 100

    if sl_pct > MAX_SL_PCT:
        print(f"  [REJECTED] {sym}: SL too wide ({sl_pct:.1f}% > {MAX_SL_PCT}%)")
        return None

    t1 = entry + sl_width * 1.0
    t2 = entry + sl_width * 2.0

    # Fetch sector for risk concentration cap
    sector = "General"
    try:
        info = yf.Ticker(ticker).info
        sector = info.get('sector', info.get('industry', 'General'))
        if not sector or sector == 'None':
            sector = "General"
    except Exception:
        sector = "General"

    print(f"  [QUALIFIED] {sym} ({sector}) | Score={relative_score:.2f} | Entry={entry:.2f} | SL={sl:.2f} ({sl_pct:.1f}%) | T2={t2:.2f}")
    return {
        "Date":          datetime.date.today().strftime("%Y-%m-%d"),
        "Stock":         sym,
        "Entry":         round(entry, 2),
        "SL":            round(sl, 2),
        "T1":            round(t1, 2),
        "T2":            round(t2, 2),
        "SL Width":      round(sl_width, 2),
        "SL %":          round(sl_pct, 2),
        "R:R T1":        1.0,
        "R:R T2":        2.0,
        "Score":         relative_score,
        "Sector":        sector
    }

# ─────────────────────────────────────────────
# EVALUATE WATCHLIST STOCK (lighter check)
# ─────────────────────────────────────────────
def evaluate_watchlist_stock(sym, bench_last, bench_df, regime_ok):
    """
    Looser evaluation for near-breakout candidates.
    Only checks: enough data, trend ok, positive 1-month return.
    Score threshold is dropped (they haven't triggered yet).
    Returns a watchlist dict or None.
    """
    ticker = sym + ".NS"
    data = yf.Ticker(ticker).history(period="1y")

    if len(data) < 50:
        return None

    data['SMA50'] = data['Close'].rolling(50).mean()
    data['Ret1m'] = (data['Close'] - data['Close'].shift(21)) / data['Close'].shift(21) * 100
    data['High52'] = data['High'].rolling(252).max() if len(data) >= 252 else data['High'].max()

    last = data.iloc[-1]
    trend_ok    = last['Close'] > last['SMA50']
    momentum_ok = data['Ret1m'].iloc[-1] > 0

    if not (trend_ok and momentum_ok and regime_ok):
        return None

    dist_from_high = ((last['High52'] - last['Close']) / last['High52']) * 100

    print(f"  [WATCHLIST] {sym} | Price={last['Close']:.2f} | {dist_from_high:.1f}% below 52W High")
    return {
        "Date":            datetime.date.today().strftime("%Y-%m-%d"),
        "Stock":           sym,
        "LTP":             round(last['Close'], 2),
        "52W High":        round(last['High52'], 2),
        "% from High":     round(dist_from_high, 2),
        "SMA50":           round(last['SMA50'], 2),
        "1M Return %":     round(data['Ret1m'].iloc[-1], 2),
        "Note":            "Near-breakout — watch for volume surge + close above high"
    }

# ─────────────────────────────────────────────
# EXCEL: SAFE NUMERIC CONVERSION & CONFIG
# ─────────────────────────────────────────────
def safe_float(val, default=0.0):
    """Safely converts val to float, handling formulas, percentages, strings, and Nones."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip()
    if val_str.startswith('='):
        return default
    val_str = val_str.replace('%', '').replace(',', '').replace('₹', '')
    try:
        return float(val_str)
    except (ValueError, TypeError):
        return default

def read_config(wb):
    capital, risk_pct, brokerage = 500000.0, 2.0, 20.0
    if 'Config' not in wb.sheetnames:
        return capital, risk_pct, brokerage
    ws = wb['Config']
    for row in ws.iter_rows(min_row=2, max_row=15, values_only=True):
        if not row[0]:
            continue
        param = str(row[0]).lower().strip()
        if "total capital" in param or param == "capital":
            capital = safe_float(row[1], 500000.0)
        elif ("max risk" in param or "risk per trade" in param) and "amount" not in param:
            risk_pct = safe_float(row[1], 2.0)
        elif "brokerage" in param:
            brokerage = safe_float(row[1], 20.0)
    return capital, risk_pct, brokerage

# ─────────────────────────────────────────────
# EXCEL: GET AVAILABLE CAPITAL
# ─────────────────────────────────────────────
def get_available_capital(wb, base_capital, brokerage):
    ws_log = wb['Trade Log']
    deployed, realized_pnl = 0.0, 0.0
    for row in range(2, ws_log.max_row + 1):
        status   = ws_log.cell(row=row, column=18).value
        qty      = safe_float(ws_log.cell(row=row, column=11).value, 0.0)
        entry    = safe_float(ws_log.cell(row=row, column=3).value, 0.0)
        exit_pr  = safe_float(ws_log.cell(row=row, column=19).value, 0.0)
        pos_size = safe_float(ws_log.cell(row=row, column=12).value, 0.0)
        if status == "OPEN":
            deployed += pos_size
        elif status and "CLOSED" in str(status).upper():
            realized_pnl += (exit_pr - entry) * qty - (2 * brokerage)
    available = base_capital - deployed + realized_pnl
    print(f"-> Base Capital:       ₹{base_capital:,.2f}")
    print(f"-> Deployed Capital:   ₹{deployed:,.2f}")
    print(f"-> Realized P&L:       ₹{realized_pnl:,.2f}")
    print(f"-> Available (Buying): ₹{available:,.2f}")
    return available

# ─────────────────────────────────────────────
# EXCEL: WRITE TRADE LOG
# ─────────────────────────────────────────────
def write_trade_log(wb, trades, base_capital, risk_pct, brokerage):
    if not trades:
        print("-> No valid trade setups today. Trade Log update skipped.")
        return

    ws = wb['Trade Log']
    today_str = datetime.date.today().strftime("%Y-%m-%d")

    # Clear today's rows to prevent duplicates
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == today_str:
            for c in range(1, 23):
                ws.cell(row=r, column=c).value = None

    # Find next empty row
    next_row = 2
    while ws.cell(row=next_row, column=1).value is not None:
        next_row += 1

    for t in trades:
        risk_amt = base_capital * (risk_pct / 100)
        qty      = math.floor(risk_amt / t['SL Width']) if t['SL Width'] > 0 else 0
        pos_size = qty * t['Entry']

        row_data = [
            t['Date'], t['Stock'], t['Entry'], t['SL'], t['T1'], t['T2'],
            t['SL %'] / 100.0, t['R:R T1'], t['R:R T2'], "GO",
            qty, round(pos_size, 2), round(risk_amt, 2),
            round(t['SL'] * 1.001, 2), t['SL'],
            round(t['T2'] * 0.999, 2), t['T2'],
            None, None, None, None, t.get('Sector', 'General')
        ]
        for col_idx, val in enumerate(row_data, 1):
            if val is not None:
                cell = ws.cell(row=next_row, column=col_idx, value=val)
                if col_idx == 7:
                    cell.number_format = '0.00%'

        print(f"-> Wrote trade: {t['Stock']} ({t.get('Sector', 'General')}) to row {next_row}")
        next_row += 1

# ─────────────────────────────────────────────
# EXCEL: WRITE WATCHLIST SHEET
# ─────────────────────────────────────────────
def write_watchlist(wb, watchlist_stocks):
    # Create or clear Watchlist sheet
    if 'Watchlist' in wb.sheetnames:
        ws = wb['Watchlist']
        ws.delete_rows(1, ws.max_row + 1)
    else:
        ws = wb.create_sheet('Watchlist')

    headers = ["Date", "Stock", "LTP (₹)", "52W High (₹)", "% from High", "SMA50", "1M Return %", "Note"]

    # Style header row
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="00F0FF")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    if not watchlist_stocks:
        ws.cell(row=2, column=1, value="No near-breakout candidates found today.")
        print("-> No watchlist candidates today.")
    else:
        for row_idx, w in enumerate(watchlist_stocks, 2):
            row_data = [
                w['Date'], w['Stock'], w['LTP'], w['52W High'],
                w['% from High'], w['SMA50'], w['1M Return %'], w['Note']
            ]
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        print(f"-> Wrote {len(watchlist_stocks)} stocks to Watchlist sheet.")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print("=" * 55)
    print("   STOCK SCREENER V2 — IMPROVED AUTOMATION ENGINE")
    print("=" * 55)
    print(f"   Date: {datetime.date.today().strftime('%d %b %Y')}")
    print(f"   Score Threshold: >{SCORE_THRESHOLD} | Max SL: {MAX_SL_PCT}% | Max Trades: {MAX_TRADES}")
    print("=" * 55)

    # ── 1. Read queries ──
    if not os.path.exists(QUERY_FILE):
        print(f"Error: {QUERY_FILE} not found.")
        return
    if not os.path.exists(WATCHLIST_QUERY_FILE):
        print(f"Warning: {WATCHLIST_QUERY_FILE} not found — watchlist disabled.")

    with open(QUERY_FILE, "r") as f:
        trade_query = f.read()

    watchlist_query = None
    if os.path.exists(WATCHLIST_QUERY_FILE):
        with open(WATCHLIST_QUERY_FILE, "r") as f:
            watchlist_query = f.read()

    # ── 2. Fetch candidates from Chartink ──
    trade_symbols     = get_chartink_stocks(trade_query, label="confirmed breakout candidates")
    watchlist_symbols = get_chartink_stocks(watchlist_query, label="near-breakout watchlist") if watchlist_query else []

    # ── 3. Load Excel config ──
    if not os.path.exists(TRACKER_FILE):
        print(f"Error: {TRACKER_FILE} not found.")
        return
    wb = openpyxl.load_workbook(TRACKER_FILE)
    base_capital, risk_pct, brokerage = read_config(wb)
    available_capital = get_available_capital(wb, base_capital, brokerage)

    # ── 4. Benchmark (fetched once) ──
    bench_last, bench_df, regime_ok = get_benchmark()
    if bench_last is None:
        return
    print(f"-> Market Regime: {'BULL (above weekly 200 SMA)' if regime_ok else 'BEAR (below weekly 200 SMA) — defensive mode'}")

    # ── 5. Evaluate confirmed breakout candidates ──
    print(f"\n── Evaluating {len(trade_symbols)} confirmed breakout candidates ──")
    qualified = []
    for sym in trade_symbols:
        t = evaluate_stock(sym, bench_last, bench_df, regime_ok)
        if t:
            qualified.append(t)

    # Sort by momentum score descending
    qualified.sort(key=lambda x: x["Score"], reverse=True)

    # Allocate by available capital & sector concentration cap (Max 2 per sector)
    MAX_SECTOR_POSITIONS = 2
    sector_counts = {}
    
    # Count open positions per sector from existing Trade Log
    ws_log = wb['Trade Log']
    for r in range(2, ws_log.max_row + 1):
        st_val = ws_log.cell(row=r, column=18).value
        sec_val = ws_log.cell(row=r, column=22).value
        if st_val == "OPEN" and sec_val:
            sec_name = str(sec_val)
            sector_counts[sec_name] = sector_counts.get(sec_name, 0) + 1

    final_trades = []
    temp_avail = available_capital
    for t in qualified:
        sec = t.get("Sector", "General")
        curr_sec_count = sector_counts.get(sec, 0)
        
        if curr_sec_count >= MAX_SECTOR_POSITIONS and sec != "General":
            print(f"  [SKIPPED SECTOR CAP] {t['Stock']} ({sec}) | Already {curr_sec_count} position(s) in this sector (Max: {MAX_SECTOR_POSITIONS})")
            continue

        risk_amt = base_capital * (risk_pct / 100)
        qty      = math.floor(risk_amt / t['SL Width']) if t['SL Width'] > 0 else 0
        pos_size = qty * t['Entry']
        if qty > 0 and pos_size <= temp_avail:
            t['Qty']           = qty
            t['Position Size'] = round(pos_size, 2)
            t['Risk Amt']      = round(risk_amt, 2)
            final_trades.append(t)
            sector_counts[sec] = curr_sec_count + 1
            temp_avail -= pos_size
            print(f"  [ALLOCATED] {t['Stock']} ({sec}) | Size=₹{pos_size:,.0f} | Remaining=₹{temp_avail:,.0f}")
            if len(final_trades) == MAX_TRADES:
                break
        else:
            print(f"  [SKIPPED CAPITAL] {t['Stock']} | Size=₹{pos_size:,.0f} > Available=₹{temp_avail:,.0f}")

    print(f"\n-> {len(final_trades)} trade(s) allocated for tomorrow.")

    # ── 6. Evaluate watchlist candidates ──
    watchlist_data = []
    if watchlist_symbols:
        # Remove any symbol already in the trade list (no duplication)
        trade_syms_set = {t['Stock'] for t in final_trades}
        unique_watchlist = [s for s in watchlist_symbols if s not in trade_syms_set]
        print(f"\n── Evaluating {len(unique_watchlist)} near-breakout watchlist candidates ──")
        for sym in unique_watchlist:
            w = evaluate_watchlist_stock(sym, bench_last, bench_df, regime_ok)
            if w:
                watchlist_data.append(w)
        # Sort by closeness to the 52W high
        watchlist_data.sort(key=lambda x: x['% from High'])
        print(f"-> {len(watchlist_data)} stock(s) added to watchlist.")

    # ── 7. Write to Excel ──
    write_trade_log(wb, final_trades, base_capital, risk_pct, brokerage)
    write_watchlist(wb, watchlist_data)

    try:
        wb.save(TRACKER_FILE)
        print("\n-> Excel Tracker saved successfully.")
    except PermissionError:
        print(f"\n[ERROR] Cannot write to Excel. Please close '{TRACKER_FILE}' and retry.")

    # ── 8. Send Telegram Alert ──
    try:
        from telegram_bot import send_daily_picks_alert
        send_daily_picks_alert(final_trades, watchlist_data)
        print("-> Telegram picks summary notification sent.")
    except Exception as e:
        print(f"-> Could not send Telegram alert: {e}")

    # ── 9. Summary ──
    print("\n" + "=" * 55)
    print(f"   DONE — {len(final_trades)} trade(s) | {len(watchlist_data)} watchlist stock(s)")
    print("=" * 55)


if __name__ == "__main__":
    main()
